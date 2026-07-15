"""Trade report — identical data pipeline to the Orders UI.

Fetches open positions from exchanges (same as AccountStreamWorker) and
closed positions via REST (same as ExchangeOrdersService).  Closed positions
are cached in data/closed_positions_cache.json so restarts are fast.

Usage:
    .venv\\Scripts\\python.exe scripts/trade_report.py [--refresh] [--last N]

    --refresh   force re-fetch closed positions from exchanges (ignores cache)
    --last N    show only the N most recently closed trades (+ summary)

Writes:
  data/trade_report.xlsx  — human table (parent = arb, children = legs)
  data/trade_report.json  — agent-friendly analysis payload (use this for trade review)
"""
from __future__ import annotations

import asyncio
import json
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

# ── project path ──────────────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_ROOT / "src"))

from dotenv import load_dotenv  # noqa: E402
load_dotenv(_ROOT / ".env")

from arbitrator.config.settings import Settings  # noqa: E402
from arbitrator.exchanges.factory import Factory  # noqa: E402
from arbitrator.domain.account.position_leg import PositionLeg  # noqa: E402
from arbitrator.domain.account.closed_position_leg import ClosedPositionLeg  # noqa: E402

_CACHE_PATH = _ROOT / "src" / "arbitrator" / "data" / "closed_positions_cache.json"
_SEEN_PATH  = _ROOT / "src" / "arbitrator" / "data" / "seen_symbols.json"
_REPORT_PATH = _ROOT / "src" / "arbitrator" / "data" / "trade_report.xlsx"
_AGENT_REPORT_PATH = _ROOT / "src" / "arbitrator" / "data" / "trade_report.json"
_CACHE_TTL_SEC = 300  # 5 min — re-fetch if older


# ── cache helpers ─────────────────────────────────────────────────────────────

def _load_cache() -> tuple[list[dict], float]:
    try:
        raw = json.loads(_CACHE_PATH.read_text(encoding="utf-8"))
        return raw["legs"], raw["fetched_at"]
    except Exception:
        return [], 0.0


def _save_cache(legs: list[ClosedPositionLeg]) -> None:
    data = {
        "fetched_at": time.time(),
        "legs": [leg.model_dump(mode="json") for leg in legs],
    }
    _CACHE_PATH.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")


def _load_seen_symbols() -> set[str]:
    try:
        return set(json.loads(_SEEN_PATH.read_text(encoding="utf-8")))
    except Exception:
        return set()


# ── fetch open positions (same as AccountStreamWorker.read_positions) ─────────

async def _fetch_open(settings: Settings, factory: Factory) -> list[PositionLeg]:
    legs: list[PositionLeg] = []
    for eid in settings.enabled_exchanges:
        if settings.credentials_for(eid) is None:
            continue
        named = factory.create(eid)
        try:
            fetched = await named.gateway.fetch_open_positions()
            legs.extend(fetched)
        except Exception as e:
            print(f"  [warn] open positions {eid}: {e}")
        finally:
            try:
                await named.gateway.close()
            except Exception:
                pass
    return legs


# ── fetch closed positions (same as ExchangeOrdersService._async_fetch_closed) ─

async def _fetch_closed(
    settings: Settings,
    factory: Factory,
    since_ms: int,
    known_symbols: list[str],
) -> list[ClosedPositionLeg]:
    legs: list[ClosedPositionLeg] = []
    for eid in settings.enabled_exchanges:
        if settings.credentials_for(eid) is None:
            continue
        named = factory.create(eid)
        try:
            fetched = await named.gateway.fetch_closed_positions(
                since_ms=since_ms, symbols=known_symbols
            )
            legs.extend(fetched)
            print(f"  {eid}: {len(fetched)} closed positions")
        except Exception as e:
            print(f"  [warn] closed positions {eid}: {e}")
        finally:
            try:
                await named.gateway.close()
            except Exception:
                pass
    return legs


# ── grouping (identical to ExchangeOrdersService._group_legs) ────────────────

def _group(
    open_legs: list[PositionLeg],
    closed_legs: list[ClosedPositionLeg],
) -> list[dict]:
    groups: list[dict] = []

    # open pairs
    by_sym: dict[str, list[PositionLeg]] = {}
    for l in open_legs:
        by_sym.setdefault(l.symbol, []).append(l)

    used_open: set[int] = set()
    for sym, legs in by_sym.items():
        shorts = [l for l in legs if l.side == "short"]
        longs  = [l for l in legs if l.side == "long"]
        for s in shorts:
            for lo in longs:
                if s.exchange_id != lo.exchange_id and id(s) not in used_open:
                    used_open.add(id(s)); used_open.add(id(lo))
                    groups.append(_open_pair(s, lo))
        for l in legs:
            if id(l) not in used_open:
                groups.append(_open_single(l))

    # closed pairs — by arb_marker_id first, then by symbol
    by_marker: dict[str, list[ClosedPositionLeg]] = {}
    by_sym_c:  dict[str, list[ClosedPositionLeg]] = {}
    for l in closed_legs:
        if l.arb_marker_id:
            by_marker.setdefault(l.arb_marker_id, []).append(l)
        else:
            by_sym_c.setdefault(l.symbol, []).append(l)

    used_closed: set[int] = set()
    for _, mlegs in by_marker.items():
        ss = [l for l in mlegs if l.side == "short"]
        ll = [l for l in mlegs if l.side == "long"]
        if ss and ll:
            used_closed.add(id(ss[0])); used_closed.add(id(ll[0]))
            groups.append(_closed_pair(ss[0], ll[0]))
        else:
            for l in mlegs:
                if id(l) not in used_closed:
                    used_closed.add(id(l))
                    groups.append(_closed_single(l))

    for _, slegs in by_sym_c.items():
        ss = [l for l in slegs if l.side == "short" and id(l) not in used_closed]
        ll = [l for l in slegs if l.side == "long"  and id(l) not in used_closed]
        for s in list(ss):
            for lo in list(ll):
                if s.exchange_id != lo.exchange_id:
                    used_closed.add(id(s)); used_closed.add(id(lo))
                    ss.remove(s); ll.remove(lo)
                    groups.append(_closed_pair(s, lo))
                    break
        for l in slegs:
            if id(l) not in used_closed:
                groups.append(_closed_single(l))

    return groups


def _open_leg(l: PositionLeg) -> dict:
    coins = l.contracts * l.contract_size
    fees = (l.opening_fee or 0.0) + (l.estimated_close_fee or 0.0)
    return {
        "exchange_id": l.exchange_id,
        "side": l.side,
        "contracts": l.contracts,
        "coins": coins,
        "volume_usdt": round(coins * l.entry_price, 2),
        "entry_price": l.entry_price,
        "exit_price": l.mark_price,
        "fees_usdt": round(fees, 4),
        "funding_usdt": round(l.accrued_funding or 0.0, 4),
        "pnl_usdt": round(l.unrealized_pnl or 0.0, 4),
    }


def _closed_leg(l: ClosedPositionLeg) -> dict:
    csize = l.contract_size if l.contract_size > 0 else 1.0
    contracts = l.contracts or 0.0
    coins = contracts * csize
    ref = l.entry_price or l.exit_price or 0.0
    return {
        "exchange_id": l.exchange_id,
        "side": l.side,
        "contracts": contracts,
        "coins": coins,
        "volume_usdt": round(coins * ref, 2),
        "entry_price": l.entry_price,
        "exit_price": l.exit_price,
        "fees_usdt": round(l.commission or 0.0, 4),
        "funding_usdt": round(l.funding or 0.0, 4),
        "pnl_usdt": round(l.realized_pnl or 0.0, 4),
    }


def _open_pair(s: PositionLeg, lo: PositionLeg) -> dict:
    pnl  = (s.unrealized_pnl or 0.0) + (lo.unrealized_pnl or 0.0)
    fees = (s.opening_fee or 0.0) + (s.estimated_close_fee or 0.0) + \
           (lo.opening_fee or 0.0) + (lo.estimated_close_fee or 0.0)
    fund = (s.accrued_funding or 0.0) + (lo.accrued_funding or 0.0)
    notional = (s.contracts * s.contract_size * s.entry_price +
                lo.contracts * lo.contract_size * lo.entry_price)
    s_mark  = s.mark_price  or s.entry_price
    lo_mark = lo.mark_price or lo.entry_price
    spread  = round((s_mark - lo_mark) / lo_mark * 100, 4) if lo_mark else None
    entry_spread = round((s.entry_price - lo.entry_price) / lo.entry_price * 100, 4) if lo.entry_price else None
    return {
        "status": "open",
        "symbol": s.symbol,
        "short_ex": s.exchange_id,
        "long_ex": lo.exchange_id,
        "opened_at": min(s.opened_at, lo.opened_at),
        "closed_at": None,
        "duration": None,
        "notional_usdt": round(notional, 2),
        "entry_spread_pct": entry_spread,
        "exit_spread_pct": None,
        "current_spread_pct": spread,
        "pnl_usdt": round(pnl, 4),
        "fees_usdt": round(fees, 4),
        "funding_usdt": round(fund, 4),
        "net_pnl_usdt": round(pnl - fees + fund, 4),
        "short_contracts": s.contracts,
        "long_contracts": lo.contracts,
        "short_entry": s.entry_price,
        "long_entry": lo.entry_price,
        "short_mark": s.mark_price,
        "long_mark": lo.mark_price,
        "legs": [_open_leg(s), _open_leg(lo)],
    }


def _open_single(l: PositionLeg) -> dict:
    notional = l.contracts * l.contract_size * l.entry_price
    fees = (l.opening_fee or 0.0) + (l.estimated_close_fee or 0.0)
    return {
        "status": "open",
        "symbol": l.symbol,
        "short_ex": l.exchange_id if l.side == "short" else None,
        "long_ex":  l.exchange_id if l.side == "long"  else None,
        "opened_at": l.opened_at,
        "closed_at": None,
        "duration": None,
        "notional_usdt": round(notional, 2),
        "entry_spread_pct": None,
        "exit_spread_pct": None,
        "current_spread_pct": None,
        "pnl_usdt": round(l.unrealized_pnl or 0.0, 4),
        "fees_usdt": round(fees, 4),
        "funding_usdt": round(l.accrued_funding or 0.0, 4),
        "net_pnl_usdt": round((l.unrealized_pnl or 0.0) - fees + (l.accrued_funding or 0.0), 4),
        "side": l.side,
        "contracts": l.contracts,
        "entry_price": l.entry_price,
        "mark_price": l.mark_price,
        "legs": [_open_leg(l)],
    }


def _closed_pair(s: ClosedPositionLeg, lo: ClosedPositionLeg) -> dict:
    pnl  = (s.realized_pnl or 0.0) + (lo.realized_pnl or 0.0)
    fees = (s.commission or 0.0) + (lo.commission or 0.0)
    fund = (s.funding or 0.0) + (lo.funding or 0.0)
    opened_at = s.opened_at or lo.opened_at
    closed_at = max(s.closed_at, lo.closed_at)
    dur = None
    if opened_at:
        secs = int((closed_at - opened_at).total_seconds())
        dur  = f"{secs//3600}h {(secs%3600)//60}m" if secs >= 3600 else f"{secs//60}m {secs%60}s"
    sx = s.exit_price; lx = lo.exit_price
    exit_spread = round((sx - lx) / lx * 100, 4) if sx and lx and lx > 0 else None
    entry_spread = (
        round((s.entry_price - lo.entry_price) / lo.entry_price * 100, 4)
        if s.entry_price and lo.entry_price and lo.entry_price > 0 else None
    )
    s_cs = s.contract_size if s.contract_size > 0 else 1.0
    lo_cs = lo.contract_size if lo.contract_size > 0 else 1.0
    s_vol = (s.contracts or 0.0) * s_cs * (s.entry_price or sx or 0.0)
    lo_vol = (lo.contracts or 0.0) * lo_cs * (lo.entry_price or lx or 0.0)
    notional = round(s_vol + lo_vol, 2) or None
    return {
        "status": "closed",
        "symbol": s.symbol,
        "short_ex": s.exchange_id,
        "long_ex": lo.exchange_id,
        "opened_at": opened_at,
        "closed_at": closed_at,
        "duration": dur,
        "notional_usdt": notional,
        "entry_spread_pct": entry_spread,
        "exit_spread_pct": exit_spread,
        "current_spread_pct": None,
        "pnl_usdt": round(pnl, 4),
        "fees_usdt": round(fees, 4),
        "funding_usdt": round(fund, 4),
        "net_pnl_usdt": round(pnl - fees + fund, 4),
        "short_entry": s.entry_price,
        "long_entry": lo.entry_price,
        "short_exit": s.exit_price,
        "long_exit": lo.exit_price,
        "short_contracts": s.contracts,
        "long_contracts": lo.contracts,
        "legs": [_closed_leg(s), _closed_leg(lo)],
    }


def _closed_single(l: ClosedPositionLeg) -> dict:
    csize = l.contract_size if l.contract_size > 0 else 1.0
    coins = (l.contracts or 0.0) * csize
    ref = l.entry_price or l.exit_price or 0.0
    notional = round(coins * ref, 2) or None
    opened_at = l.opened_at
    closed_at = l.closed_at
    dur = None
    if opened_at:
        secs = int((closed_at - opened_at).total_seconds())
        dur  = f"{secs//3600}h {(secs%3600)//60}m" if secs >= 3600 else f"{secs//60}m {secs%60}s"
    return {
        "status": "closed",
        "symbol": l.symbol,
        "short_ex": l.exchange_id if l.side == "short" else None,
        "long_ex":  l.exchange_id if l.side == "long"  else None,
        "opened_at": opened_at,
        "closed_at": closed_at,
        "duration": dur,
        "notional_usdt": notional,
        "entry_spread_pct": None,
        "exit_spread_pct": None,
        "current_spread_pct": None,
        "pnl_usdt": round(l.realized_pnl or 0.0, 4),
        "fees_usdt": round(l.commission or 0.0, 4),
        "funding_usdt": round(l.funding or 0.0, 4),
        "net_pnl_usdt": round((l.realized_pnl or 0.0) - (l.commission or 0.0) + (l.funding or 0.0), 4),
        "side": l.side,
        "contracts": l.contracts,
        "entry_price": l.entry_price,
        "exit_price": l.exit_price,
        "legs": [_closed_leg(l)],
    }


# ── rendering ─────────────────────────────────────────────────────────────────

W = 72

def _fmt(dt: datetime | None) -> str:
    if dt is None:
        return "—"
    local = dt.astimezone() if dt.tzinfo else dt
    return local.strftime("%m/%d %H:%M")


def _fmt_xlsx(dt: datetime | None) -> str:
    if dt is None:
        return ""
    local = dt.astimezone() if dt.tzinfo else dt
    return local.strftime("%d %H:%M")


def _pair(g: dict) -> str:
    return f"{g.get('short_ex') or '?'}->{g.get('long_ex') or '?'}"


def _sym(g: dict) -> str:
    return g["symbol"].replace("/USDT:USDT", "")


def _pct(v: object) -> str:
    return f"{v:+.4f}%" if isinstance(v, (int, float)) else "—"


def _n(v: object, d: int = 4) -> str:
    return f"{v:.{d}f}" if isinstance(v, (int, float)) else "—"


_XLSX_COLS = (
    "#", "Asset", "Leg", "Open", "Close", "Dur",
    "USDT", "Coins", "Entry,In%", "Exit,Out%",
    "Fee", "Fund", "PnL", "NET",
)
_XLSX_WIDTHS = (
    3.0, 7.0, 9.5, 7.0, 7.0, 5.5,
    5.5, 5.5, 11.0, 11.0,
    5.0, 5.0, 5.5, 5.5,
)


def _cell_entry_in(price: object, spr: object) -> str:
    p = f"{price:.6g}" if isinstance(price, (int, float)) else ""
    s = f"{spr:+.4f}%" if isinstance(spr, (int, float)) else ""
    if p and s:
        return f"{p} {s}"
    return p or s or ""


def _cell_exit_out(price: object, spr: object) -> str:
    return _cell_entry_in(price, spr)


def _xlsx_parent_row(g: dict, idx: int | None = None) -> list[object]:
    coins = sum((leg.get("coins") or 0.0) for leg in g.get("legs") or [])
    exit_or_cur = g.get("current_spread_pct") if g["status"] == "open" else g.get("exit_spread_pct")
    s = (g.get("short_ex") or "?")[:4]
    lo = (g.get("long_ex") or "?")[:4]
    return [
        idx if idx is not None else "",
        _sym(g),
        f"S·{s} L·{lo}",
        _fmt_xlsx(g["opened_at"]),
        _fmt_xlsx(g["closed_at"]),
        g.get("duration") or "",
        g.get("notional_usdt"),
        round(coins, 2) if coins else None,
        _cell_entry_in(None, g.get("entry_spread_pct")),
        _cell_exit_out(None, exit_or_cur),
        g["fees_usdt"],
        g["funding_usdt"],
        g["pnl_usdt"],
        g["net_pnl_usdt"],
    ]


def _xlsx_leg_row(leg: dict) -> list[object]:
    side = (leg.get("side") or "?")[:1]
    ex = str(leg.get("exchange_id") or "")[:5]
    return [
        "",
        "",
        f"└{ex} {side}",
        "",
        "",
        "",
        leg.get("volume_usdt"),
        round(float(leg["coins"]), 2) if leg.get("coins") is not None else None,
        _cell_entry_in(leg.get("entry_price"), None),
        _cell_exit_out(leg.get("exit_price"), None),
        leg.get("fees_usdt"),
        leg.get("funding_usdt") or 0.0,
        leg.get("pnl_usdt") or 0.0,
        None,
    ]


def _write_xlsx(
    open_g: list[dict],
    shown_closed: list[dict],
    closed_total: int,
    last_n: int | None,
    total_open_pnl: float,
    total_closed_pnl: float,
    total_fees: float,
    total_funding: float,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "trades"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    closed_label = f"{closed_total}" + (
        f" (last {len(shown_closed)})" if last_n is not None else ""
    )
    ws.append([f"Trade report — {now}"])
    ws.append(["Open", "Closed", "Unrealized", "Realized NET", "Fees", "Funding"])
    ws.append([
        len(open_g), closed_label, total_open_pnl, total_closed_pnl, total_fees, total_funding,
    ])
    ws.append([])

    base = Font(name="Calibri", size=8)
    parent_font = Font(name="Calibri", size=8, bold=True)
    head_fill = PatternFill("solid", fgColor="1A1F29")
    head_font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    parent_fill = PatternFill("solid", fgColor="E8ECF2")
    leg_fill = PatternFill("solid", fgColor="F7F8FA")
    compact = Alignment(vertical="center", wrap_text=False)

    def write_block(title: str, groups: list[dict], numbered: bool) -> None:
        if not groups:
            return
        ws.append([title])
        ws.append(list(_XLSX_COLS))
        head_row = ws.max_row
        for cell in ws[head_row]:
            cell.fill = head_fill
            cell.font = head_font
        for i, g in enumerate(groups, 1):
            ws.append(_xlsx_parent_row(g, i if numbered else None))
            prow = ws.max_row
            for cell in ws[prow]:
                cell.font = parent_font
                cell.fill = parent_fill
            net_cell = ws.cell(prow, 14)
            net_v = g["net_pnl_usdt"]
            net_cell.font = Font(
                name="Calibri", size=8, bold=True,
                color=("1B7F4E" if net_v > 0 else "C0392B" if net_v < 0 else "000000"),
            )
            for leg in g.get("legs") or []:
                ws.append(_xlsx_leg_row(leg))
                for cell in ws[ws.max_row]:
                    cell.fill = leg_fill
                    cell.font = base
        ws.append([])

    if last_n is None:
        write_block("Open positions", open_g, False)
    closed_title = f"Last {len(shown_closed)} closed" if last_n is not None else "Closed trades"
    write_block(closed_title, shown_closed, True)

    if last_n is not None and shown_closed:
        nets = [g["net_pnl_usdt"] for g in shown_closed]
        wins = sum(1 for n in nets if n > 0)
        ws.append([f"Last {len(shown_closed)} summary"])
        ws.append(["Wins", "Losses", "Net", "Avg", "Best", "Worst", "Fees", "Realized"])
        ws.append([
            wins, len(nets) - wins, sum(nets), sum(nets) / len(nets),
            max(nets), min(nets),
            sum(g["fees_usdt"] for g in shown_closed),
            sum(g["pnl_usdt"] for g in shown_closed),
        ])

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 12
        for cell in row:
            f = cell.font
            cell.font = Font(
                name="Calibri",
                size=8,
                bold=bool(f.bold) if f else False,
                color=f.color if f and f.color else None,
            )
            cell.alignment = compact
    for i, width in enumerate(_XLSX_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width
    ws.freeze_panes = "A5"

    wb.save(_REPORT_PATH)
    for stale in (_REPORT_PATH.with_suffix(".md"), _REPORT_PATH.with_suffix(".html")):
        if stale.exists():
            stale.unlink()


def _iso(dt: object) -> str | None:
    if isinstance(dt, datetime):
        return dt.isoformat()
    return None


def _agent_group(g: dict) -> dict[str, object]:
    legs_out: list[dict[str, object]] = []
    for leg in g.get("legs") or []:
        legs_out.append({
            "exchange_id": leg.get("exchange_id"),
            "side": leg.get("side"),
            "coins": leg.get("coins"),
            "contracts": leg.get("contracts"),
            "usdt": leg.get("volume_usdt"),
            "entry_price": leg.get("entry_price"),
            "exit_price": leg.get("exit_price"),
            "fees_usdt": leg.get("fees_usdt"),
            "funding_usdt": leg.get("funding_usdt"),
            "pnl_usdt": leg.get("pnl_usdt"),
        })
    return {
        "status": g["status"],
        "symbol": g["symbol"],
        "asset": _sym(g),
        "short_exchange_id": g.get("short_ex"),
        "long_exchange_id": g.get("long_ex"),
        "opened_at": _iso(g.get("opened_at")),
        "closed_at": _iso(g.get("closed_at")),
        "duration": g.get("duration"),
        "usdt": g.get("notional_usdt"),
        "coins": round(sum((leg.get("coins") or 0.0) for leg in g.get("legs") or []), 4),
        "entry_spread_pct": g.get("entry_spread_pct"),
        "exit_spread_pct": g.get("exit_spread_pct"),
        "current_spread_pct": g.get("current_spread_pct"),
        "fees_usdt": g["fees_usdt"],
        "funding_usdt": g["funding_usdt"],
        "pnl_usdt": g["pnl_usdt"],
        "net_pnl_usdt": g["net_pnl_usdt"],
        "legs": legs_out,
    }


def _write_agent_json(
    open_g: list[dict],
    shown_closed: list[dict],
    closed_total: int,
    last_n: int | None,
    total_open_pnl: float,
    total_closed_pnl: float,
    total_fees: float,
    total_funding: float,
) -> None:
    nets = [g["net_pnl_usdt"] for g in shown_closed]
    wins = sum(1 for n in nets if n > 0) if nets else 0
    payload: dict[str, object] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "scope": {
            "last_n": last_n,
            "open_count": len(open_g),
            "closed_total": closed_total,
            "closed_shown": len(shown_closed),
        },
        "totals": {
            "unrealized_pnl_usdt": round(total_open_pnl, 4),
            "realized_net_pnl_usdt": round(total_closed_pnl, 4),
            "fees_usdt": round(total_fees, 4),
            "funding_usdt": round(total_funding, 4),
        },
        "shown_summary": {
            "wins": wins,
            "losses": len(nets) - wins,
            "net_pnl_usdt": round(sum(nets), 4) if nets else 0.0,
            "avg_net_pnl_usdt": round(sum(nets) / len(nets), 4) if nets else 0.0,
            "best_net_pnl_usdt": max(nets) if nets else None,
            "worst_net_pnl_usdt": min(nets) if nets else None,
            "fees_usdt": round(sum(g["fees_usdt"] for g in shown_closed), 4),
            "realized_pnl_usdt": round(sum(g["pnl_usdt"] for g in shown_closed), 4),
        } if shown_closed else None,
        "open": [_agent_group(g) for g in open_g] if last_n is None else [],
        "closed": [_agent_group(g) for g in shown_closed],
    }
    _AGENT_REPORT_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )
def _render(groups: list[dict], last_n: int | None = None) -> None:
    open_g    = [g for g in groups if g["status"] == "open"]
    closed_g  = [g for g in groups if g["status"] == "closed"]
    closed_g.sort(
        key=lambda g: g["closed_at"] or datetime.min.replace(tzinfo=timezone.utc),
        reverse=True,
    )
    shown_closed = closed_g[:last_n] if last_n is not None else closed_g

    total_open_pnl   = sum(g["pnl_usdt"] for g in open_g)
    total_closed_pnl = sum(g["net_pnl_usdt"] for g in closed_g)
    total_fees       = sum(g["fees_usdt"] for g in groups)
    total_funding    = sum(g["funding_usdt"] for g in groups)

    print("=" * W)
    print(f"  TRADE REPORT  --  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * W)
    print(f"  Open positions :  {len(open_g)}")
    print(f"  Closed trades  :  {len(closed_g)}"
          + (f"  (showing last {len(shown_closed)})" if last_n is not None else ""))
    print(f"  Unrealized PnL :  {total_open_pnl:+.4f} USDT")
    print(f"  Realized PnL   :  {total_closed_pnl:+.4f} USDT  (after fees+funding)")
    print(f"  Total fees     :  {total_fees:.4f} USDT")
    print(f"  Total funding  :  {total_funding:+.4f} USDT")
    print("=" * W)

    if open_g and last_n is None:
        print("\n-- OPEN POSITIONS " + "-" * (W - 18))
        for g in open_g:
            pair  = _pair(g)
            s_ex  = g.get("short_ex") or "?"
            l_ex  = g.get("long_ex")  or "?"
            notio = f"{g['notional_usdt']} USDT" if g["notional_usdt"] else "?"
            sprd  = f"spread={g['current_spread_pct']:+.4f}%" if g.get("current_spread_pct") is not None else ""
            entry = f"entry={g.get('entry_spread_pct'):+.4f}%" if g.get("entry_spread_pct") is not None else ""
            imb   = ""
            if g.get("short_contracts") and g.get("long_contracts"):
                ratio = g["short_contracts"] / g["long_contracts"]
                if abs(ratio - 1.0) > 0.05:
                    imb = f"  ! imbalance {ratio:.2f}x"
            print(f"\n  {_sym(g):<10} {pair:<22} {notio:<14} since {_fmt(g['opened_at'])}")
            print(f"             {entry:<24} {sprd}")
            print(f"             unrealized={g['pnl_usdt']:+.4f}  fees={g['fees_usdt']:.4f}  "
                  f"funding={g['funding_usdt']:+.4f}  net={g['net_pnl_usdt']:+.4f} USDT{imb}")
            if g.get("short_contracts"):
                print(f"             short {s_ex}: {g['short_contracts']} @ {g.get('short_entry')}  mark={g.get('short_mark')}")
            if g.get("long_contracts"):
                print(f"             long  {l_ex}: {g['long_contracts']} @ {g.get('long_entry')}  mark={g.get('long_mark')}")

    if shown_closed:
        title = f"LAST {len(shown_closed)} CLOSED" if last_n is not None else "CLOSED TRADES"
        print(f"\n-- {title} " + "-" * max(0, W - 4 - len(title)))
        for g in shown_closed:
            pair = _pair(g)
            s_ex = g.get("short_ex") or "?"
            l_ex = g.get("long_ex")  or "?"
            notio = f"{g['notional_usdt']} USDT" if g["notional_usdt"] else "?"
            sprd = f"exit_spread={g['exit_spread_pct']:+.4f}%" if g.get("exit_spread_pct") is not None else ""
            net  = g["net_pnl_usdt"]
            sign = "+" if net >= 0 else ""
            print(f"\n  {_sym(g):<10} {pair:<22} {notio:<14} dur={g.get('duration') or '?'}")
            print(f"             opened={_fmt(g['opened_at'])}  closed={_fmt(g['closed_at'])}  {sprd}")
            print(f"             realized={g['pnl_usdt']:+.4f}  fees={g['fees_usdt']:.4f}  "
                  f"funding={g['funding_usdt']:+.4f}  NET={sign}{net:.4f} USDT")
            if g.get("short_entry"):
                print(f"             short {s_ex}: entry={g['short_entry']}  exit={g.get('short_exit')}")
            if g.get("long_entry"):
                print(f"             long  {l_ex}: entry={g['long_entry']}  exit={g.get('long_exit')}")

        if last_n is not None:
            nets = [g["net_pnl_usdt"] for g in shown_closed]
            wins = sum(1 for n in nets if n > 0)
            print(f"\n-- LAST {len(shown_closed)} SUMMARY " + "-" * max(0, W - 18 - len(str(len(shown_closed)))))
            print(f"  wins={wins}  losses={len(nets) - wins}  "
                  f"net={sum(nets):+.4f}  avg={sum(nets)/len(nets):+.4f}  "
                  f"best={max(nets):+.4f}  worst={min(nets):+.4f}")
            print(f"  fees={sum(g['fees_usdt'] for g in shown_closed):.4f}  "
                  f"realized={sum(g['pnl_usdt'] for g in shown_closed):+.4f}")

    _write_xlsx(
        open_g, shown_closed, len(closed_g), last_n,
        total_open_pnl, total_closed_pnl, total_fees, total_funding,
    )
    _write_agent_json(
        open_g, shown_closed, len(closed_g), last_n,
        total_open_pnl, total_closed_pnl, total_fees, total_funding,
    )
    print(f"\n  wrote {_REPORT_PATH}")
    print(f"  wrote {_AGENT_REPORT_PATH}")
    print("\n" + "=" * W)


# ── main ──────────────────────────────────────────────────────────────────────

async def _main(force_refresh: bool, last_n: int | None) -> None:
    settings = Settings()  # type: ignore[call-arg]
    factory  = Factory(settings)

    print("Fetching open positions from exchanges...")
    open_legs = await _fetch_open(settings, factory)
    print(f"  total open: {len(open_legs)}")

    open_symbols = {l.symbol for l in open_legs}
    seen_symbols = _load_seen_symbols() | open_symbols
    known_symbols = sorted(seen_symbols)

    cached_legs, fetched_at = _load_cache()
    age = time.time() - fetched_at
    if force_refresh or age > _CACHE_TTL_SEC or not cached_legs:
        print(f"\nFetching closed positions from exchanges (cache age={int(age)}s)...")
        since_ms = int(time.time() * 1000) - 7 * 24 * 3600 * 1000
        closed_legs = await _fetch_closed(settings, factory, since_ms, known_symbols)
        _save_cache(closed_legs)
        print(f"  cached {len(closed_legs)} closed legs to {_CACHE_PATH.name}")
    else:
        print(f"\nUsing cached closed positions (age={int(age)}s, {len(cached_legs)} legs)")
        closed_legs = [ClosedPositionLeg(**r) for r in cached_legs]

    print()
    groups = _group(open_legs, closed_legs)
    _render(groups, last_n=last_n)


def _parse_last(argv: list[str]) -> int | None:
    if "--last" not in argv:
        return None
    i = argv.index("--last")
    if i + 1 >= len(argv) or not argv[i + 1].isdigit():
        print("usage: --last N  (positive integer)", file=sys.stderr)
        sys.exit(2)
    n = int(argv[i + 1])
    if n < 1:
        print("usage: --last N  (positive integer)", file=sys.stderr)
        sys.exit(2)
    return n


if __name__ == "__main__":
    force = "--refresh" in sys.argv
    asyncio.run(_main(force, _parse_last(sys.argv)))
