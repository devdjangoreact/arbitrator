import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.styles import Font  # type: ignore


class ExcelTradesLogger:
    def __init__(self, directory: str = "logs/trades") -> None:
        self.directory = directory
        os.makedirs(self.directory, exist_ok=True)
        self.headers = [
            "Time", "Tick MS", "Rank", "Symbol", "Short Ex", "Long Ex", "Cache Spread %",
            "Threshold %", "Pool", "Max Pos", "Dup Sym", "Cooldown",
            "Gateways", "Cache Thr", "Check 1", "Check 2", "Exec",
            "Fresh Spread %", "Est Fill %", "Notional", "Outcome", "Detail",
            "Short Book", "Long Book"
        ]
        self._font = Font(size=8)

    def _get_filename(self) -> str:
        date_str = datetime.now().strftime("%Y-%m-%d")
        return os.path.join(self.directory, f"open_candidates_{date_str}.xlsx")

    def _ensure_workbook(self, filename: str) -> Any:
        if os.path.exists(filename):
            try:
                return load_workbook(filename)
            except Exception:
                pass
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = "Open Candidates"
            ws.append(self.headers)
            # Set font for headers
            for cell in ws[1]:
                cell.font = self._font
            # Set minimal column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = min(adjusted_width, 15) # Keep it narrow
        return wb

    def log_candidate(self, trace: Any) -> None:
        filename = self._get_filename()
        wb = self._ensure_workbook(filename)
        ws = wb.active

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        def _stage(name: str) -> str:
            return str(trace.stages.get(name, "—"))

        chk = trace.check2 if trace.check2 and not trace.check2.passed else trace.check1
        if chk is None and trace.check1 is not None:
            chk = trace.check1

        fresh = round(chk.fresh_spread, 3) if chk and chk.fresh_spread is not None else None
        est = round(chk.estimated_spread, 3) if chk and chk.estimated_spread is not None else None
        notional = round(chk.notional, 1) if chk and chk.notional is not None else None

        outcome = trace.final_outcome

        short_book = ""
        long_book = ""
        if chk and (chk.short_book or chk.long_book):
            short_book = str(chk.short_book)
            long_book = str(chk.long_book)

        row = [
            now_str,
            trace.tick_ms,
            trace.rank,
            trace.symbol,
            trace.short_ex,
            trace.long_ex,
            round(trace.cache_spread_pct, 3),
            trace.threshold_pct,
            _stage("pool"),
            _stage("max_pos"),
            _stage("dup_sym"),
            _stage("cooldown"),
            _stage("gateways"),
            _stage("cache_thr"),
            _stage("check1"),
            _stage("check2"),
            _stage("execute"),
            fresh,
            est,
            notional,
            outcome,
            trace.final_detail,
            short_book,
            long_book
        ]
        if ws:
            ws.append(row)
            # Apply font to the new row
            for cell in ws[ws.max_row]:
                cell.font = self._font
        wb.save(filename)
